import uuid
from datetime import date
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest

from app.core.exceptions import DomainError
from app.services.liquidaciones import LiquidacionService


def _make_mock_result(rows: list | None = None, scalar=None):
    m = MagicMock()
    if rows is not None:
        m.unique().scalars().all.return_value = rows
    if scalar is not None:
        m.scalar_one.return_value = scalar
    return m


class TestCalcularLiquidacion:
    @pytest.fixture
    def service(self):
        return LiquidacionService(tenant_id=uuid.uuid4())

    @pytest.fixture
    def mock_session(self):
        return AsyncMock()

    @pytest.mark.asyncio
    async def test_calcular_periodo_cerrado_raise_error(self, service, mock_session):
        service._repo.periodo_cerrado = AsyncMock(return_value=True)
        with pytest.raises(DomainError, match="ya está cerrado"):
            await service.calcular(mock_session, uuid.uuid4(), "2026-06")

    @pytest.mark.asyncio
    async def test_calcular_sin_comisiones_retorna_vacio(self, service, mock_session):
        service._repo.periodo_cerrado = AsyncMock(return_value=False)
        service._repo.soft_delete_abiertas = AsyncMock()
        mock_result = _make_mock_result(rows=[])

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        result = await service.calcular(
            mock_session, uuid.uuid4(), "2026-06"
        )
        assert result == []

    @pytest.mark.asyncio
    async def test_calcular_reemplaza_el_borrador_antes_de_insertar(
        self, service, mock_session
    ):
        """PA recalculo: antes de insertar las nuevas liquidaciones, calcular()
        debe descartar (soft-delete) las 'Abierta' previas de esa cohorte+
        período, para no duplicar filas al recalcular un período abierto."""
        service._repo.periodo_cerrado = AsyncMock(return_value=False)
        service._repo.soft_delete_abiertas = AsyncMock()
        mock_result = _make_mock_result(rows=[])

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        cohorte_id = uuid.uuid4()

        await service.calcular(mock_session, cohorte_id, "2026-06")

        service._repo.soft_delete_abiertas.assert_awaited_once_with(
            mock_session, cohorte_id, "2026-06"
        )

    @pytest.mark.asyncio
    async def test_calcular_periodo_cerrado_no_borra_nada(self, service, mock_session):
        """Si el período ya está cerrado, calcular() debe abortar con 409
        ANTES de tocar las liquidaciones existentes — nunca se debe llamar
        a soft_delete_abiertas en ese caso."""
        service._repo.periodo_cerrado = AsyncMock(return_value=True)
        service._repo.soft_delete_abiertas = AsyncMock()

        with pytest.raises(DomainError, match="ya está cerrado"):
            await service.calcular(mock_session, uuid.uuid4(), "2026-06")

        service._repo.soft_delete_abiertas.assert_not_awaited()

    @pytest.mark.asyncio
    async def test_calcular_acumula_plus_por_clave_de_materia_segun_cantidad_comisiones(
        self, service, mock_session
    ):
        """PA-22/PA-23: el Plus se calcula por materia_id -> clave_plus, NO por
        las letras de comisión ("A", "B"). Un docente con 1 asignación a una
        materia con clave_plus="PROG" y 2 comisiones (["A", "B"]) debe acumular
        2x el Plus(PROG, rol), tal cual RN-33/RN-34."""
        service._repo.periodo_cerrado = AsyncMock(return_value=False)
        service._repo.soft_delete_abiertas = AsyncMock()

        usuario_id = uuid.uuid4()
        materia_id = uuid.uuid4()

        asignacion = MagicMock()
        asignacion.usuario_id = usuario_id
        asignacion.rol = "PROFESOR"
        asignacion.materia_id = materia_id
        asignacion.comisiones = ["A", "B"]

        salario_base = MagicMock(monto=1000.0)
        salario_plus = MagicMock(monto=200.0)
        user = MagicMock(facturador=False)

        results = [
            _make_mock_result(rows=[asignacion]),  # select(Asignacion)
        ]
        mock_session.execute = AsyncMock(side_effect=results + [
            MagicMock(**{"unique.return_value.scalar_one_or_none.return_value": salario_base}),
            MagicMock(scalar_one_or_none=MagicMock(return_value="PROG")),
            MagicMock(**{"unique.return_value.scalar_one_or_none.return_value": salario_plus}),
            MagicMock(**{"unique.return_value.scalar_one_or_none.return_value": user}),
        ])

        captured = {}

        async def fake_create(session, data):
            captured.update(data)
            return SimpleNamespace(id=uuid.uuid4(), **data)

        service._repo.create = AsyncMock(side_effect=fake_create)

        result = await service.calcular(mock_session, uuid.uuid4(), "2026-06")

        assert len(result) == 1
        assert captured["monto_base"] == 1000.0
        assert captured["monto_plus"] == 400.0  # 200 * 2 comisiones
        assert result[0]["total"] == 1400.0

    @pytest.mark.asyncio
    async def test_calcular_sin_clave_plus_no_suma_plus(self, service, mock_session):
        """Una materia sin clave_plus asignada (None) no debe sumar Plus —
        evita que se intente buscar SalarioPlus con una clave inexistente."""
        service._repo.periodo_cerrado = AsyncMock(return_value=False)
        service._repo.soft_delete_abiertas = AsyncMock()

        usuario_id = uuid.uuid4()
        materia_id = uuid.uuid4()

        asignacion = MagicMock()
        asignacion.usuario_id = usuario_id
        asignacion.rol = "TUTOR"
        asignacion.materia_id = materia_id
        asignacion.comisiones = ["A"]

        salario_base = MagicMock(monto=500.0)
        user = MagicMock(facturador=False)

        mock_session.execute = AsyncMock(side_effect=[
            _make_mock_result(rows=[asignacion]),
            MagicMock(**{"unique.return_value.scalar_one_or_none.return_value": salario_base}),
            MagicMock(scalar_one_or_none=MagicMock(return_value=None)),  # sin clave_plus
            MagicMock(**{"unique.return_value.scalar_one_or_none.return_value": user}),
        ])

        captured = {}

        async def fake_create(session, data):
            captured.update(data)
            return SimpleNamespace(id=uuid.uuid4(), **data)

        service._repo.create = AsyncMock(side_effect=fake_create)

        result = await service.calcular(mock_session, uuid.uuid4(), "2026-06")

        assert len(result) == 1
        assert captured["monto_plus"] == 0.0
        assert result[0]["total"] == 500.0

    @pytest.mark.asyncio
    async def test_get_salario_base_vigente_none(self, service, mock_session):
        mock_result = MagicMock()
        mock_result.unique().scalar_one_or_none.return_value = None

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        result = await service._get_salario_base_vigente(
            mock_session, "PROFESOR", date(2026, 6, 1)
        )
        assert result is None

    @pytest.mark.asyncio
    async def test_get_clave_plus_none_si_materia_sin_clave(self, service, mock_session):
        mock_result = MagicMock()
        mock_result.scalar_one_or_none.return_value = None

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        result = await service._get_clave_plus(mock_session, uuid.uuid4())
        assert result is None

    @pytest.mark.asyncio
    async def test_get_clave_plus_devuelve_clave_de_la_materia(self, service, mock_session):
        mock_result = MagicMock()
        mock_result.scalar_one_or_none.return_value = "PROG"

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        result = await service._get_clave_plus(mock_session, uuid.uuid4())
        assert result == "PROG"

    @pytest.mark.asyncio
    async def test_get_salario_plus_vigente_none(self, service, mock_session):
        mock_result = MagicMock()
        mock_result.unique().scalar_one_or_none.return_value = None

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        result = await service._get_salario_plus_vigente(
            mock_session, "PROG", "PROFESOR", date(2026, 6, 1)
        )
        assert result is None

    @pytest.mark.asyncio
    async def test_cerrar_liquidacion(self, service, mock_session):
        mock_liq = MagicMock()
        mock_liq.id = uuid.uuid4()
        mock_liq.estado = "Cerrada"
        service._repo.cerrar = AsyncMock(return_value=mock_liq)
        result = await service.cerrar(mock_session, mock_liq.id)
        assert result["estado"] == "Cerrada"

    @pytest.mark.asyncio
    async def test_cerrar_liquidacion_ya_cerrada(self, service, mock_session):
        service._repo.cerrar = AsyncMock(
            side_effect=DomainError(
                detail="La liquidación ya está cerrada",
                context={},
            )
        )
        with pytest.raises(DomainError, match="ya está cerrada"):
            await service.cerrar(mock_session, uuid.uuid4())

    @pytest.mark.asyncio
    async def test_listar_por_periodo(self, service, mock_session):
        usuario_id = uuid.uuid4()
        mock_liq = MagicMock()
        mock_liq.id = uuid.uuid4()
        mock_liq.cohorte_id = uuid.uuid4()
        mock_liq.periodo = "2026-06"
        mock_liq.usuario_id = usuario_id
        mock_liq.rol = "PROFESOR"
        mock_liq.comisiones = ["PROG"]
        mock_liq.monto_base = 500.0
        mock_liq.monto_plus = 200.0
        mock_liq.total = 700.0
        mock_liq.es_nexo = False
        mock_liq.excluido_por_factura = False
        mock_liq.estado = "Abierta"
        service._repo.listar_por_periodo = AsyncMock(return_value=[mock_liq])

        user = MagicMock()
        user.id = usuario_id
        user.nombre = "Juan"
        user.apellidos = "Pérez"
        user.email = "juan@demo.local"
        mock_session.execute = AsyncMock(
            return_value=MagicMock(**{"unique.return_value.scalars.return_value.all.return_value": [user]})
        )

        result = await service.listar(mock_session, "2026-06")
        assert len(result) == 1
        assert result[0]["total"] == 700.0
        assert result[0]["docente_nombre"] == "Juan Pérez"

    @pytest.mark.asyncio
    async def test_listar_usuario_sin_resolver_usa_id_como_fallback(self, service, mock_session):
        """Si el usuario_id no aparece en el mapa resuelto (ej. usuario borrado),
        no debe romper — cae al string del id en vez de crashear."""
        usuario_id = uuid.uuid4()
        mock_liq = MagicMock()
        mock_liq.id = uuid.uuid4()
        mock_liq.cohorte_id = uuid.uuid4()
        mock_liq.periodo = "2026-06"
        mock_liq.usuario_id = usuario_id
        mock_liq.rol = "PROFESOR"
        mock_liq.comisiones = []
        mock_liq.monto_base = 500.0
        mock_liq.monto_plus = 0.0
        mock_liq.total = 500.0
        mock_liq.es_nexo = False
        mock_liq.excluido_por_factura = False
        mock_liq.estado = "Abierta"
        service._repo.listar_por_periodo = AsyncMock(return_value=[mock_liq])

        mock_session.execute = AsyncMock(
            return_value=MagicMock(**{"unique.return_value.scalars.return_value.all.return_value": []})
        )

        result = await service.listar(mock_session, "2026-06")
        assert result[0]["docente_nombre"] == str(usuario_id)


class TestHistorialLiquidacion:
    @pytest.fixture
    def service(self):
        return LiquidacionService(tenant_id=uuid.uuid4())

    @pytest.fixture
    def mock_session(self):
        return AsyncMock()

    def _make_liq(self, periodo, cohorte_id, total, estado="Abierta"):
        liq = MagicMock()
        liq.periodo = periodo
        liq.cohorte_id = cohorte_id
        liq.total = total
        liq.estado = estado
        return liq

    @pytest.mark.asyncio
    async def test_historial_agrupa_por_periodo_y_cohorte(self, service, mock_session):
        cohorte_id = uuid.uuid4()
        liqs = [
            self._make_liq("2026-05", cohorte_id, 1000.0),
            self._make_liq("2026-05", cohorte_id, 500.0),
            self._make_liq("2026-06", cohorte_id, 700.0),
        ]
        service._repo.listar_todas = AsyncMock(return_value=liqs)

        result = await service.historial(mock_session)

        assert len(result) == 2
        mayo = next(r for r in result if r["periodo"] == "2026-05")
        assert mayo["total_docentes"] == 2
        assert mayo["monto_total"] == 1500.0

    @pytest.mark.asyncio
    async def test_historial_estado_cerrada_solo_si_todas_cerradas(self, service, mock_session):
        cohorte_id = uuid.uuid4()
        liqs = [
            self._make_liq("2026-05", cohorte_id, 1000.0, estado="Cerrada"),
            self._make_liq("2026-05", cohorte_id, 500.0, estado="Abierta"),
        ]
        service._repo.listar_todas = AsyncMock(return_value=liqs)

        result = await service.historial(mock_session)

        assert result[0]["estado"] == "Abierta"

    @pytest.mark.asyncio
    async def test_historial_filtra_por_desde_hasta_y_estado(self, service, mock_session):
        cohorte_id = uuid.uuid4()
        liqs = [
            self._make_liq("2026-01", cohorte_id, 100.0, estado="Cerrada"),
            self._make_liq("2026-05", cohorte_id, 200.0, estado="Abierta"),
            self._make_liq("2026-08", cohorte_id, 300.0, estado="Abierta"),
        ]
        service._repo.listar_todas = AsyncMock(return_value=liqs)

        result = await service.historial(mock_session, desde="2026-02", hasta="2026-07")

        assert len(result) == 1
        assert result[0]["periodo"] == "2026-05"

    @pytest.mark.asyncio
    async def test_historial_sin_liquidaciones_devuelve_vacio(self, service, mock_session):
        service._repo.listar_todas = AsyncMock(return_value=[])
        result = await service.historial(mock_session)
        assert result == []

    @pytest.mark.asyncio
    async def test_exportar_planilla_genera_xlsx_con_filas_correctas(
        self, service, mock_session
    ):
        import openpyxl
        from io import BytesIO

        usuario_id = uuid.uuid4()
        liq = MagicMock()
        liq.usuario_id = usuario_id
        liq.rol = "PROFESOR"
        liq.comisiones = ["A", "B"]
        liq.monto_base = 1000.0
        liq.monto_plus = 600.0
        liq.total = 1600.0
        liq.es_nexo = False
        liq.excluido_por_factura = False
        liq.estado = "Abierta"
        service._repo.listar_por_periodo = AsyncMock(return_value=[liq])

        user = MagicMock()
        user.id = usuario_id
        user.nombre = "Juan"
        user.apellidos = "Pérez"
        user.email = "juan@demo.local"
        mock_session.execute = AsyncMock(
            return_value=MagicMock(**{"unique.return_value.scalars.return_value.all.return_value": [user]})
        )

        xlsx_bytes = await service.exportar_planilla(mock_session, "2026-06")

        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == [
            "Docente", "Email", "Rol", "Comisiones",
            "Base", "Plus", "Total", "NEXO", "Facturante", "Estado",
        ]
        row = [c.value for c in ws[2]]
        assert row[0] == "Juan Pérez"
        assert row[1] == "juan@demo.local"
        assert row[4] == 1000.0
        assert row[5] == 600.0
        assert row[6] == 1600.0

    @pytest.mark.asyncio
    async def test_exportar_planilla_usuario_sin_nombre_usa_email(
        self, service, mock_session
    ):
        """Usuarios creados sin nombre/apellidos (ej. admin de bootstrap) no
        deben mostrar "None None" en la planilla — se usa el email."""
        import openpyxl
        from io import BytesIO

        usuario_id = uuid.uuid4()
        liq = MagicMock()
        liq.usuario_id = usuario_id
        liq.rol = "ADMIN"
        liq.comisiones = None
        liq.monto_base = 0.0
        liq.monto_plus = 0.0
        liq.total = 0.0
        liq.es_nexo = False
        liq.excluido_por_factura = False
        liq.estado = "Abierta"
        service._repo.listar_por_periodo = AsyncMock(return_value=[liq])

        user = MagicMock()
        user.id = usuario_id
        user.nombre = None
        user.apellidos = None
        user.email = "admin@demo.local"
        mock_session.execute = AsyncMock(
            return_value=MagicMock(**{"unique.return_value.scalars.return_value.all.return_value": [user]})
        )

        xlsx_bytes = await service.exportar_planilla(mock_session, "2026-06")

        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        row = [c.value for c in ws[2]]
        assert row[0] == "admin@demo.local"

    @pytest.mark.asyncio
    async def test_exportar_planilla_sin_liquidaciones_genera_xlsx_solo_headers(
        self, service, mock_session
    ):
        import openpyxl
        from io import BytesIO

        service._repo.listar_por_periodo = AsyncMock(return_value=[])

        xlsx_bytes = await service.exportar_planilla(mock_session, "2026-06")

        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.max_row == 1

    @pytest.mark.asyncio
    async def test_obtener_kpis(self, service, mock_session):
        service._repo.obtener_kpis = AsyncMock(
            return_value={
                "total_general": 1000.0,
                "total_nexo": 500.0,
                "total_facturas_pendientes": 0.0,
                "total_facturas_abonadas": 0.0,
                "cantidad_docentes_general": 2,
                "cantidad_docentes_nexo": 1,
                "cantidad_docentes_facturantes": 0,
            }
        )
        mock_result = MagicMock()
        mock_result.scalar_one.return_value = 0

        async def execute_side(*args, **kwargs):
            return mock_result

        mock_session.execute.side_effect = execute_side
        kpis = await service.obtener_kpis(mock_session, "2026-06")
        assert kpis["total_general"] == 1000.0
        assert kpis["cantidad_docentes_general"] == 2
