import uuid
from datetime import date

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import DomainError
from app.models.asignacion import Asignacion
from app.models.materia import Materia
from app.models.salario_base import SalarioBase
from app.models.salario_plus import SalarioPlus
from app.models.user import User
from app.repositories.liquidacion import LiquidacionRepository


class LiquidacionService:
    def __init__(self, tenant_id: uuid.UUID) -> None:
        self._tenant_id = tenant_id
        self._repo = LiquidacionRepository(tenant_id)

    async def calcular(
        self,
        session: AsyncSession,
        cohorte_id: uuid.UUID,
        periodo: str,
    ) -> list[dict]:
        if await self._repo.periodo_cerrado(session, cohorte_id, periodo):
            raise DomainError(
                detail=f"El período {periodo} ya está cerrado para este cohorte",
                context={"cohorte_id": str(cohorte_id), "periodo": periodo},
            )

        await self._repo.soft_delete_abiertas(session, cohorte_id, periodo)

        anio, mes = periodo.split("-")
        fecha_liq = date(int(anio), int(mes), 1)

        result = await session.execute(
            select(Asignacion).where(
                Asignacion.tenant_id == self._tenant_id,
                Asignacion.cohorte_id == cohorte_id,
                Asignacion.desde <= fecha_liq,
                (
                    (Asignacion.hasta.is_(None))
                    | (Asignacion.hasta > fecha_liq)
                ),
                Asignacion.deleted_at.is_(None),
            )
        )
        asignaciones = list(result.unique().scalars().all())

        docentes_map: dict[tuple[uuid.UUID, str], dict] = {}
        for a in asignaciones:
            key = (a.usuario_id, a.rol)
            if key not in docentes_map:
                docentes_map[key] = {
                    "usuario_id": a.usuario_id,
                    "rol": a.rol,
                    "comisiones": [],
                    "materias_comisiones": [],
                }
            if a.comisiones:
                docentes_map[key]["comisiones"].extend(a.comisiones)
                if a.materia_id is not None:
                    docentes_map[key]["materias_comisiones"].append(
                        (a.materia_id, len(a.comisiones))
                    )

        if not docentes_map:
            return []

        liquidaciones: list[dict] = []
        for key, docente in docentes_map.items():
            usuario_id, rol = key

            salario_base = await self._get_salario_base_vigente(
                session, rol, fecha_liq
            )
            monto_base = float(salario_base.monto) if salario_base else 0.0

            comisiones = docente["comisiones"]
            monto_plus = 0.0
            for materia_id, cantidad_comisiones in docente["materias_comisiones"]:
                clave_plus = await self._get_clave_plus(session, materia_id)
                if clave_plus is None:
                    continue
                salario_plus = await self._get_salario_plus_vigente(
                    session, clave_plus, rol, fecha_liq
                )
                if salario_plus:
                    monto_plus += float(salario_plus.monto) * cantidad_comisiones

            total = monto_base + monto_plus
            es_nexo = rol == "NEXO"

            user_result = await session.execute(
                select(User).where(User.id == usuario_id, User.tenant_id == self._tenant_id)
            )
            user = user_result.unique().scalar_one_or_none()
            excluido = user.facturador if user else False

            liquidacion_data = {
                "cohorte_id": cohorte_id,
                "periodo": periodo,
                "usuario_id": usuario_id,
                "rol": rol,
                "comisiones": list(set(comisiones)),
                "monto_base": monto_base,
                "monto_plus": monto_plus,
                "total": total,
                "es_nexo": es_nexo,
                "excluido_por_factura": excluido,
                "estado": "Abierta",
            }
            liq = await self._repo.create(session, liquidacion_data)
            liquidaciones.append({
                "id": str(liq.id),
                "usuario_id": str(liq.usuario_id),
                "rol": liq.rol,
                "monto_base": float(liq.monto_base),
                "monto_plus": float(liq.monto_plus),
                "total": float(liq.total),
                "es_nexo": liq.es_nexo,
                "excluido_por_factura": liq.excluido_por_factura,
                "estado": liq.estado,
            })

        return liquidaciones

    async def _get_salario_base_vigente(
        self,
        session: AsyncSession,
        rol: str,
        fecha: date,
    ) -> SalarioBase | None:
        result = await session.execute(
            select(SalarioBase).where(
                SalarioBase.tenant_id == self._tenant_id,
                SalarioBase.rol == rol,
                SalarioBase.desde <= fecha,
                (
                    (SalarioBase.hasta.is_(None))
                    | (SalarioBase.hasta > fecha)
                ),
                SalarioBase.deleted_at.is_(None),
            ).order_by(SalarioBase.desde.desc()).limit(1)
        )
        return result.unique().scalar_one_or_none()

    async def _get_clave_plus(
        self,
        session: AsyncSession,
        materia_id: uuid.UUID,
    ) -> str | None:
        result = await session.execute(
            select(Materia.clave_plus).where(
                Materia.id == materia_id,
                Materia.tenant_id == self._tenant_id,
            )
        )
        return result.scalar_one_or_none()

    async def _get_salario_plus_vigente(
        self,
        session: AsyncSession,
        grupo: str,
        rol: str,
        fecha: date,
    ) -> SalarioPlus | None:
        result = await session.execute(
            select(SalarioPlus).where(
                SalarioPlus.tenant_id == self._tenant_id,
                SalarioPlus.grupo == grupo,
                SalarioPlus.rol == rol,
                SalarioPlus.desde <= fecha,
                (
                    (SalarioPlus.hasta.is_(None))
                    | (SalarioPlus.hasta > fecha)
                ),
                SalarioPlus.deleted_at.is_(None),
            ).order_by(SalarioPlus.desde.desc()).limit(1)
        )
        return result.unique().scalar_one_or_none()

    async def listar(
        self,
        session: AsyncSession,
        periodo: str,
        cohorte_id: uuid.UUID | None = None,
    ) -> list[dict]:
        items = await self._repo.listar_por_periodo(session, periodo, cohorte_id)
        return [
            {
                "id": str(l.id),
                "cohorte_id": str(l.cohorte_id),
                "periodo": l.periodo,
                "usuario_id": str(l.usuario_id),
                "rol": l.rol,
                "comisiones": l.comisiones,
                "monto_base": float(l.monto_base),
                "monto_plus": float(l.monto_plus),
                "total": float(l.total),
                "es_nexo": l.es_nexo,
                "excluido_por_factura": l.excluido_por_factura,
                "estado": l.estado,
            }
            for l in items
        ]

    async def exportar_planilla(
        self,
        session: AsyncSession,
        periodo: str,
        cohorte_id: uuid.UUID | None = None,
    ) -> bytes:
        items = await self._repo.listar_por_periodo(session, periodo, cohorte_id)

        usuario_ids = {item.usuario_id for item in items}
        usuarios: dict[uuid.UUID, User] = {}
        if usuario_ids:
            result = await session.execute(
                select(User).where(
                    User.id.in_(usuario_ids),
                    User.tenant_id == self._tenant_id,
                )
            )
            usuarios = {u.id: u for u in result.unique().scalars().all()}

        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Liquidaciones {periodo}"

        headers = [
            "Docente", "Email", "Rol", "Comisiones",
            "Base", "Plus", "Total", "NEXO", "Facturante", "Estado",
        ]
        header_font = Font(bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font

        for i, item in enumerate(items, 2):
            usuario = usuarios.get(item.usuario_id)
            nombre = self._formatear_nombre(usuario) if usuario else str(item.usuario_id)
            email = usuario.email if usuario else ""
            ws.cell(row=i, column=1, value=nombre)
            ws.cell(row=i, column=2, value=email)
            ws.cell(row=i, column=3, value=item.rol)
            ws.cell(row=i, column=4, value=", ".join(item.comisiones) if item.comisiones else "")
            ws.cell(row=i, column=5, value=float(item.monto_base))
            ws.cell(row=i, column=6, value=float(item.monto_plus))
            ws.cell(row=i, column=7, value=float(item.total))
            ws.cell(row=i, column=8, value="Sí" if item.es_nexo else "No")
            ws.cell(row=i, column=9, value="Sí" if item.excluido_por_factura else "No")
            ws.cell(row=i, column=10, value=item.estado)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["D"].width = 20

        from io import BytesIO

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _formatear_nombre(usuario: User) -> str:
        partes = [p for p in (usuario.nombre, usuario.apellidos) if p]
        return " ".join(partes) if partes else usuario.email

    async def cerrar(
        self,
        session: AsyncSession,
        liquidacion_id: uuid.UUID,
    ) -> dict:
        liq = await self._repo.cerrar(session, liquidacion_id)
        return {
            "id": str(liq.id),
            "estado": liq.estado,
        }

    async def obtener_kpis(
        self,
        session: AsyncSession,
        periodo: str,
    ) -> dict:
        from app.models.factura import Factura
        from sqlalchemy import select, func

        kpis = await self._repo.obtener_kpis(session, periodo)

        facturas_query = select(
            func.coalesce(func.sum(Factura.tamano_kb), 0)
        ).where(
            Factura.tenant_id == self._tenant_id,
            Factura.periodo == periodo,
            Factura.estado == "Pendiente",
        )
        result = await session.execute(facturas_query)
        kpis["total_facturas_pendientes"] = float(result.scalar_one())

        facturas_abonadas_query = select(
            func.coalesce(func.sum(Factura.tamano_kb), 0)
        ).where(
            Factura.tenant_id == self._tenant_id,
            Factura.periodo == periodo,
            Factura.estado == "Abonada",
        )
        result = await session.execute(facturas_abonadas_query)
        kpis["total_facturas_abonadas"] = float(result.scalar_one())

        return kpis
