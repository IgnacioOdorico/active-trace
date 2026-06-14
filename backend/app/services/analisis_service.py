import io
import uuid
from collections import defaultdict
from datetime import datetime
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import DomainError
from app.repositories.calificacion_repository import CalificacionRepository
from app.repositories.padron_repository import (
    EntradaPadronRepository,
    VersionPadronRepository,
)
from app.repositories.umbral_repository import UmbralRepository
from app.services.audit_service import AuditLogService


class AnalisisService:
    def __init__(self, tenant_id: uuid.UUID) -> None:
        self._calificacion_repo = CalificacionRepository(tenant_id)
        self._umbral_repo = UmbralRepository(tenant_id)
        self._version_repo = VersionPadronRepository(tenant_id)
        self._entrada_repo = EntradaPadronRepository(tenant_id)
        self._audit_service = AuditLogService(tenant_id)

    async def atrasados(
        self,
        db: AsyncSession,
        materia_id: uuid.UUID,
    ) -> list[dict[str, Any]]:
        versiones = await self._version_repo.get_active_by_materia(db, materia_id)
        if not versiones:
            raise DomainError("No hay un padrón activo para esta materia")

        version_ids = [v.id for v in versiones]
        todas_entradas = []
        for vid in version_ids:
            entradas = await self._entrada_repo.get_by_version(db, vid)
            todas_entradas.extend(entradas)

        calificaciones = await self._calificacion_repo.get_by_materia(db, materia_id)

        umbral = await self._umbral_repo.get_by_asignacion(db, uuid.UUID(int=0))
        umbral_pct, _ = self._umbral_repo.get_umbral_efectivo(umbral)

        calif_por_alumno: dict[uuid.UUID, list] = defaultdict(list)
        for c in calificaciones:
            calif_por_alumno[c.entrada_padron_id].append(c)

        atrasados_list: list[dict[str, Any]] = []

        for entrada in todas_entradas:
            problemas: list[dict[str, str]] = []
            alumno_califs = calif_por_alumno.get(entrada.id, [])

            if not alumno_califs:
                problemas.append({
                    "nombre_actividad": "General",
                    "motivo": "actividades_faltantes",
                })
            else:
                for c in alumno_califs:
                    if (
                        c.nota_numerica is not None
                        and c.nota_numerica < umbral_pct
                    ):
                        problemas.append({
                            "nombre_actividad": c.nombre_actividad,
                            "motivo": "nota_inferior_umbral",
                        })

            if problemas:
                atrasados_list.append({
                    "entrada_padron_id": str(entrada.id),
                    "nombre": entrada.nombre,
                    "apellidos": entrada.apellidos,
                    "email": entrada.email,
                    "comision": entrada.comision,
                    "actividades_problematicas": problemas,
                })

        return atrasados_list

    async def ranking(
        self,
        db: AsyncSession,
        materia_id: uuid.UUID,
    ) -> list[dict[str, Any]]:
        ranking_data = await self._calificacion_repo.get_ranking_aprobadas(
            db, materia_id
        )

        if not ranking_data:
            return []

        entrada_ids = [item[0] for item in ranking_data]
        ranking_map: dict[uuid.UUID, int] = dict(ranking_data)

        versiones = await self._version_repo.get_active_by_materia(db, materia_id)
        todas_entradas = []
        if versiones:
            for v in versiones:
                entradas = await self._entrada_repo.get_by_version(db, v.id)
                todas_entradas.extend(entradas)

        entrada_map = {e.id: e for e in todas_entradas}

        calificaciones = await self._calificacion_repo.get_by_materia(db, materia_id)
        total_por_alumno: dict[uuid.UUID, set[str]] = defaultdict(set)
        for c in calificaciones:
            total_por_alumno[c.entrada_padron_id].add(c.nombre_actividad)

        result: list[dict[str, Any]] = []
        for eid in entrada_ids:
            entrada = entrada_map.get(eid)
            if entrada is None:
                continue
            result.append({
                "entrada_padron_id": str(eid),
                "nombre": entrada.nombre,
                "apellidos": entrada.apellidos,
                "comision": entrada.comision,
                "actividades_aprobadas": ranking_map[eid],
                "total_actividades": len(total_por_alumno.get(eid, set())),
            })

        result.sort(key=lambda x: x["actividades_aprobadas"], reverse=True)
        return result

    async def reportes_rapidos(
        self,
        db: AsyncSession,
        materia_id: uuid.UUID,
    ) -> dict[str, Any]:
        calificaciones = await self._calificacion_repo.get_by_materia(db, materia_id)

        total_calificaciones = len(calificaciones)
        if not calificaciones:
            return {
                "total_alumnos": 0,
                "total_actividades": 0,
                "total_calificaciones": 0,
                "promedio_aprobacion_general": None,
                "alumnos_atrasados_count": 0,
                "alumnos_aprobados_count": 0,
                "sin_datos": True,
            }

        actividades_set: set[str] = set()
        alumnos_set: set[uuid.UUID] = set()
        aprobados_count = 0
        for c in calificaciones:
            actividades_set.add(c.nombre_actividad)
            alumnos_set.add(c.entrada_padron_id)
            if c.aprobado:
                aprobados_count += 1

        versiones = await self._version_repo.get_active_by_materia(db, materia_id)
        total_alumnos = 0
        if versiones:
            for v in versiones:
                entradas = await self._entrada_repo.get_by_version(db, v.id)
                total_alumnos += len(entradas)

        atrasados = await self.atrasados(db, materia_id)

        return {
            "total_alumnos": total_alumnos,
            "total_actividades": len(actividades_set),
            "total_calificaciones": total_calificaciones,
            "promedio_aprobacion_general": round(
                aprobados_count / total_calificaciones, 4
            ) if total_calificaciones > 0 else None,
            "alumnos_atrasados_count": len(atrasados),
            "alumnos_aprobados_count": len(alumnos_set) - len(atrasados),
            "sin_datos": False,
        }

    async def notas_finales(
        self,
        db: AsyncSession,
        materia_id: uuid.UUID,
        actividades: list[tuple[str, str]],
    ) -> list[dict[str, Any]]:
        actividad_nombres = [a[0] for a in actividades]

        rows = await self._calificacion_repo.get_notas_por_alumno(
            db, materia_id, actividad_nombres
        )

        if not rows:
            raise DomainError(
                f"Actividad no encontrada: {actividad_nombres[0]}"
            )

        alumno_data: dict[uuid.UUID, dict[str, Any]] = {}
        for calif, entrada in rows:
            eid = calif.entrada_padron_id
            if eid not in alumno_data:
                alumno_data[eid] = {
                    "entrada_padron_id": str(eid),
                    "nombre": entrada.nombre,
                    "apellidos": entrada.apellidos,
                    "comision": entrada.comision,
                    "notas": [],
                    "nota_final": None,
                    "actividades_textuales": [],
                }

            alumno_data[eid]["notas"].append({
                "actividad": calif.nombre_actividad,
                "nota_numerica": calif.nota_numerica,
                "nota_textual": calif.nota_textual,
            })

        result = list(alumno_data.values())
        for item in result:
            numericas = [
                n["nota_numerica"]
                for n in item["notas"]
                if n["nota_numerica"] is not None
            ]
            textuales = [
                n["actividad"]
                for n in item["notas"]
                if n["nota_textual"] is not None
                and n["nota_numerica"] is None
            ]
            item["actividades_textuales"] = textuales
            if numericas:
                item["nota_final"] = round(sum(numericas) / len(numericas), 2)

        actividades_encontradas = set()
        for item in result:
            for n in item["notas"]:
                actividades_encontradas.add(n["actividad"])

        for aname in actividad_nombres:
            if aname not in actividades_encontradas:
                raise DomainError(f"Actividad no encontrada: {aname}")

        return result

    async def exportar_tps(
        self,
        db: AsyncSession,
        materia_id: uuid.UUID,
    ) -> Any:
        rows = await self._calificacion_repo.get_by_materia_con_entrada(
            db, materia_id
        )

        entregas: list[dict[str, str]] = []
        for calif, entrada in rows:
            if calif.nota_textual is not None and calif.nota_numerica is None:
                entregas.append({
                    "alumno": f"{entrada.nombre} {entrada.apellidos}",
                    "actividad": calif.nombre_actividad,
                    "valor_entrega": calif.nota_textual,
                })

        if not entregas:
            return {
                "total": 0,
                "mensaje": "No se encontraron entregas sin corregir",
            }

        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TPs sin corregir"

        headers = ["Alumno", "Actividad", "Estado"]
        header_font = Font(bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font

        for i, entrega in enumerate(entregas, 2):
            ws.cell(row=i, column=1, value=entrega["alumno"])
            ws.cell(row=i, column=2, value=entrega["actividad"])
            ws.cell(row=i, column=3, value=entrega["valor_entrega"])

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return {
            "content": output.getvalue(),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"tps_sin_corregir_{materia_id}.xlsx",
            "total": len(entregas),
        }

    async def monitor_general(
        self,
        db: AsyncSession,
        *,
        materia_id: uuid.UUID | None = None,
        regional: str | None = None,
        comision: str | None = None,
        q: str | None = None,
        estado: str | None = None,
        pagina: int = 1,
        por_pagina: int = 50,
    ) -> dict[str, Any]:
        items_raw, total = await self._calificacion_repo.get_filtrado(
            db,
            materia_id=materia_id,
            regional=regional,
            comision=comision,
            q=q,
            pagina=pagina,
            por_pagina=por_pagina,
        )

        alumno_map: dict[uuid.UUID, dict[str, Any]] = {}
        for calif, entrada in items_raw:
            eid = calif.entrada_padron_id
            if eid not in alumno_map:
                alumno_map[eid] = {
                    "entrada_padron_id": str(eid),
                    "nombre": entrada.nombre,
                    "apellidos": entrada.apellidos,
                    "email": entrada.email,
                    "comision": entrada.comision,
                    "regional": entrada.regional,
                    "materia_id": str(calif.materia_id) if calif.materia_id else None,
                    "total_actividades": 0,
                    "aprobadas": 0,
                    "estado": "al_dia",
                }
            alumno_map[eid]["total_actividades"] += 1
            if calif.aprobado:
                alumno_map[eid]["aprobadas"] += 1

        items = list(alumno_map.values())
        for item in items:
            if item["aprobadas"] < item["total_actividades"]:
                item["estado"] = "atrasado"
            else:
                item["estado"] = "al_dia"

        if estado == "atrasado":
            items = [i for i in items if i["estado"] == "atrasado"]
        elif estado == "al_dia":
            items = [i for i in items if i["estado"] == "al_dia"]

        total_paginas = max(1, (total + por_pagina - 1) // por_pagina)

        return {
            "items": items,
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": total_paginas,
        }

    async def monitor_seguimiento(
        self,
        db: AsyncSession,
        *,
        materias_ids: list[uuid.UUID] | None = None,
        comision: str | None = None,
        entrada_padron_id: uuid.UUID | None = None,
        actividad_minima: str | None = None,
        desde: datetime | None = None,
        hasta: datetime | None = None,
        pagina: int = 1,
        por_pagina: int = 50,
    ) -> dict[str, Any]:
        items_raw: list = []
        total = 0
        if materias_ids:
            for mid in materias_ids:
                sub_items, sub_total = await self._calificacion_repo.get_filtrado(
                    db,
                    materia_id=mid,
                    comision=comision,
                    entrada_padron_id=entrada_padron_id,
                    desde=desde,
                    hasta=hasta,
                    pagina=pagina,
                    por_pagina=por_pagina,
                )
                items_raw.extend(sub_items)
                total += sub_total
        else:
            items_raw, total = await self._calificacion_repo.get_filtrado(
                db,
                comision=comision,
                entrada_padron_id=entrada_padron_id,
                desde=desde,
                hasta=hasta,
                pagina=pagina,
                por_pagina=por_pagina,
            )

        alumno_map: dict[uuid.UUID, dict[str, Any]] = {}
        for calif, entrada in items_raw:
            eid = calif.entrada_padron_id
            if eid not in alumno_map:
                alumno_map[eid] = {
                    "entrada_padron_id": str(eid),
                    "nombre": entrada.nombre,
                    "apellidos": entrada.apellidos,
                    "email": entrada.email,
                    "comision": entrada.comision,
                    "regional": entrada.regional,
                    "materia_id": str(calif.materia_id) if calif.materia_id else None,
                    "total_actividades": 0,
                    "aprobadas": 0,
                    "estado": "al_dia",
                }
            alumno_map[eid]["total_actividades"] += 1
            if calif.aprobado:
                alumno_map[eid]["aprobadas"] += 1

        items = list(alumno_map.values())
        for item in items:
            if item["aprobadas"] < item["total_actividades"]:
                item["estado"] = "atrasado"
            else:
                item["estado"] = "al_dia"

        if actividad_minima and items:
            items_filtered = []
            for item in items:
                eid = uuid.UUID(item["entrada_padron_id"])
                califs = await self._calificacion_repo.get_by_materia(db, uuid.UUID(int=0))
                has_minima = any(
                    c.entrada_padron_id == eid
                    and c.nombre_actividad == actividad_minima
                    for c in califs
                )
                if has_minima:
                    items_filtered.append(item)
            items = items_filtered

        total_paginas = max(1, (total + por_pagina - 1) // por_pagina)

        return {
            "items": items,
            "total": len(items),
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": total_paginas,
        }
